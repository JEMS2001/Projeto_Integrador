from email.mime.image import MIMEImage
import pandas as pd
import io
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.chart import BarChart, Reference, PieChart, ScatterChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart import ScatterChart, Series, Reference

# Mudança na importação
from .gerador_relatorio import GeradorRelatorio
 

class GerenciadorRelatorios:
    def __init__(self, conexao_bd, enviador_email):
        self.gerador = GeradorRelatorio(conexao_bd)
        self.enviador = enviador_email

    def gerar_e_enviar_relatorios(self, periodo='semanal'):
        try:
            empresas = self.gerador.obter_empresas()
            if empresas.empty:
                print("Nenhuma empresa encontrada para gerar relatórios.")
                return
        except Exception as e:
            print(f"Erro ao obter empresas: {str(e)}")
            return
        
        if periodo == 'semanal':
            data_fim = datetime.now().date()
            data_inicio = data_fim - timedelta(days=7)
        elif periodo == 'mensal':
            data_fim = datetime.now().date()
            data_inicio = data_fim.replace(day=1)
        
        for _, empresa in empresas.iterrows():
            id_empresa = empresa['id_empresa']
            nome_empresa = empresa['nome']
            email_empresa = empresa['email']
            
            try:
                # Gerar relatórios específicos para a empresa
                df_projetos = self.gerador.relatorio_projetos(id_empresa, data_inicio, data_fim)
                df_desempenho = self.gerador.relatorio_desempenho_membros(id_empresa, data_inicio, data_fim)
                df_uso = self.gerador.relatorio_uso_sistema(id_empresa, data_inicio, data_fim)
                
                # Verificar se algum dos DataFrames está vazio
                if df_projetos.empty and df_desempenho.empty and df_uso.empty:
                    print(f"Nenhum dado encontrado para a empresa {nome_empresa}. Relatório não será gerado.")
                    continue
                
                # Criar visualizações
                graficos = self._criar_visualizacoes(df_projetos, df_desempenho, df_uso)
                
                # Criar planilha Excel com múltiplas abas
                excel_buffer = self._criar_excel([
                    ('Projetos', df_projetos),
                    ('Desempenho', df_desempenho),
                    ('Uso do Sistema', df_uso)
                ])
                
                # Criar corpo do e-mail
                corpo_email = self._criar_corpo_email(nome_empresa, periodo, data_inicio, data_fim, graficos)
                
                # Enviar e-mail
                self.enviador.enviar_email(
                    email_empresa,
                    f"Relatório {periodo.capitalize()} da Empresa {nome_empresa} - {data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}",
                    corpo_email,
                    excel_buffer,
                    graficos
                )
                
                print(f"Relatório gerado e enviado com sucesso para a empresa {nome_empresa}")
            
            except Exception as e:
                print(f"Erro ao gerar ou enviar relatório para a empresa {nome_empresa}: {str(e)}")

    def _criar_visualizacoes(self, df_projetos, df_desempenho, df_uso):
        graficos = []

        # Gráfico de barras para status dos projetos
        plt.figure(figsize=(10, 6))
        sns.countplot(x='status', data=df_projetos)
        plt.title('Status dos Projetos')
        plt.tight_layout()
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        graficos.append(('status_projetos.png', buffer.getvalue()))
        plt.close()

        # Gráfico de dispersão para tarefas concluídas vs média de dias para conclusão
        plt.figure(figsize=(10, 6))
        sns.scatterplot(x='tarefas_concluidas', y='media_dias_conclusao', data=df_desempenho)
        plt.title('Desempenho dos Membros')
        plt.xlabel('Tarefas Concluídas')
        plt.ylabel('Média de Dias para Conclusão')
        plt.tight_layout()
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        graficos.append(('desempenho_membros.png', buffer.getvalue()))
        plt.close()

        # Gráfico de barras para uso do sistema
        plt.figure(figsize=(10, 6))
        sns.barplot(x='nome', y='horas_total', data=df_uso.head(10))
        plt.title('Top 10 Usuários por Horas de Uso')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png')
        buffer.seek(0)
        graficos.append(('uso_sistema.png', buffer.getvalue()))
        plt.close()

        return graficos

    def _criar_excel(self, dataframes):
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            workbook = writer.book

            for nome, df in dataframes:
                df.to_excel(writer, sheet_name=nome, index=False)
                worksheet = writer.sheets[nome]
                
                # Apply styles
                self._aplicar_estilos_excel(worksheet, df)
                
                # Add charts
                self._adicionar_graficos_excel(workbook, worksheet, df, nome)

            # Set the first sheet as active
            workbook.active = 0

        buffer.seek(0)
        return buffer
    
    def _aplicar_estilos_excel(self, worksheet, df):
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply header styles
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Apply styles to data cells
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Auto-adjust column widths
        dim_holder = DimensionHolder(worksheet=worksheet)

        for col in range(worksheet.min_column, worksheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(worksheet, min=col, max=col, width=20)

        worksheet.column_dimensions = dim_holder

        # Freeze the header row
        worksheet.freeze_panes = 'A2'

    def _adicionar_graficos_excel(self, workbook, worksheet, df, nome):
        if nome == 'Projetos':
            self._adicionar_grafico_projetos(workbook, worksheet, df)
        elif nome == 'Desempenho':
            self._adicionar_grafico_desempenho(workbook, worksheet, df)
        elif nome == 'Uso do Sistema':
            self._adicionar_grafico_uso_sistema(workbook, worksheet, df)

    def _adicionar_grafico_projetos(self, workbook, worksheet, df):
        # Create a pie chart for project status
        pie = PieChart()
        labels = Reference(worksheet, min_col=5, min_row=2, max_row=worksheet.max_row)
        data = Reference(worksheet, min_col=6, min_row=1, max_row=worksheet.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Distribuição de Status dos Projetos"

        # Add data labels
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]

        # Add chart to the worksheet
        worksheet.add_chart(pie, "I2")

    def _adicionar_grafico_desempenho(self, workbook, worksheet, df):
        # Create a scatter chart for member performance
        scatter = ScatterChart()
        scatter.title = "Desempenho dos Membros"
        scatter.x_axis.title = "Tarefas Concluídas"
        scatter.y_axis.title = "Média de Dias para Conclusão"

        x_values = Reference(worksheet, min_col=2, min_row=2, max_row=worksheet.max_row)
        y_values = Reference(worksheet, min_col=3, min_row=2, max_row=worksheet.max_row)
        
        # Create a new series and add it to the chart
        series = Series(y_values, x_values, title_from_data=False)
        scatter.series.append(series)

        # Add chart to the worksheet
        worksheet.add_chart(scatter, "E2")


    def _adicionar_grafico_uso_sistema(self, workbook, worksheet, df):
        # Create a bar chart for system usage
        chart = BarChart()
        chart.type = "col"
        chart.title = "Top 10 Usuários por Horas de Uso"
        chart.y_axis.title = 'Horas Totais'
        chart.x_axis.title = 'Usuários'

        data = Reference(worksheet, min_col=3, min_row=1, max_row=11, max_col=3)
        cats = Reference(worksheet, min_col=1, min_row=2, max_row=11)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Add chart to the worksheet
        worksheet.add_chart(chart, "E2")


    def _criar_corpo_email(self, nome_empresa, periodo, data_inicio, data_fim, graficos):
        corpo = f"""
        <!DOCTYPE html>
        <html lang="pt-BR">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Relatório {periodo.capitalize()} - {nome_empresa}</title>
            <style>
                @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;700&display=swap');
                body {{
                    font-family: 'Roboto', Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 800px;
                    margin: 0 auto;
                    padding: 20px;
                    background-color: #f4f4f4;
                }}
                .container {{
                    background-color: #ffffff;
                    border-radius: 8px;
                    padding: 30px;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
                }}
                h1, h2, h3 {{
                    color: #2c3e50;
                }}
                h1 {{
                    font-size: 28px;
                    border-bottom: 2px solid #3498db;
                    padding-bottom: 10px;
                    margin-bottom: 20px;
                }}
                h2 {{
                    font-size: 24px;
                    margin-top: 30px;
                }}
                p {{
                    margin-bottom: 15px;
                }}
                ul {{
                    padding-left: 20px;
                }}
                li {{
                    margin-bottom: 10px;
                }}
                .periodo {{
                    font-weight: bold;
                    color: #3498db;
                }}
                .grafico {{
                    margin-top: 20px;
                    text-align: center;
                }}
                .grafico img {{
                    max-width: 100%;
                    height: auto;
                    border-radius: 4px;
                    box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
                }}
                .footer {{
                    margin-top: 30px;
                    text-align: center;
                    font-size: 14px;
                    color: #7f8c8d;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Relatório {periodo.capitalize()} - {nome_empresa}</h1>
                <p>Prezado cliente,</p>
                <p>Segue o relatório detalhado da sua empresa para o período de <span class="periodo">{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}</span>.</p>
                
                <h2>Conteúdo do Relatório</h2>
                <ul>
                    <li>Resumo de Projetos</li>
                    <li>Desempenho dos Membros</li>
                    <li>Uso do Sistema</li>
                </ul>
                
                <h2>Visualizações Principais</h2>
        """
        
        for nome, _ in graficos:
            corpo += f"""
                <div class="grafico">
                    <img src="cid:{nome}" alt="{nome.replace('.png', '').replace('_', ' ').title()}">
                </div>
            """
        
        corpo += """
                <p>Para uma análise mais detalhada, por favor, consulte o arquivo Excel anexo a este e-mail.</p>
                
                <h2>Próximos Passos</h2>
                <p>Recomendamos que você:</p>
                <ul>
                    <li>Revise o desempenho dos projetos e identifique áreas de melhoria</li>
                    <li>Analise o desempenho individual dos membros da equipe</li>
                    <li>Verifique os padrões de uso do sistema para otimizar a produtividade</li>
                </ul>
                
                <p>Caso tenha alguma dúvida ou necessite de esclarecimentos adicionais, não hesite em entrar em contato conosco.</p>
                
                <div class="footer">
                    <p>Este é um e-mail automático. Por favor, não responda diretamente a esta mensagem.</p>
                </div>
            </div>
        </body>
        </html>
        """
        return corpo
